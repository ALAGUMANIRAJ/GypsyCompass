import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.conf import settings
from pathlib import Path
import datetime
import requests

def get_user_ip_location(ip_address):
    """Get user's approximate location from IP address."""
    try:
        # Use free ipapi service
        response = requests.get(f'https://ipapi.co/{ip_address}/json/', timeout=5)
        if response.status_code == 200:
            data = response.json()
            city = data.get('city', 'Unknown')
            region = data.get('region', '')
            country = data.get('country_name', 'Unknown')
            return f"{city}, {region}, {country}".strip(', ')
    except:
        pass
    return 'Unknown'

def get_client_ip(request):
    """Extract real IP from request."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR', '127.0.0.1')
    return ip

def init_excel():
    """Initialize the Excel workbook with headers if it doesn't exist."""
    excel_path = Path(settings.USER_DATA_EXCEL)
    if excel_path.exists():
        try:
            return openpyxl.load_workbook(excel_path)
        except:
            pass # Corruption perhaps? proceed to recreate

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Travel Data"

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a6b4a", end_color="1a6b4a", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        "S.No", "User Name", "IP Address", "IP-Based Location",
        "Start Location (User)", "Budget", "Currency", "Travel Type",
        "Group Size", "Travel Scope", "No. of Days", "Food & Accommodation",
        "Travel Medium", "Destination Styles", "Timestamp"
    ]

    col_widths = [6, 20, 18, 30, 25, 12, 10, 12, 12, 18, 12, 22, 15, 40, 25]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 30
    wb.save(excel_path)
    return wb

def save_user_data(user_data, ip_address, ip_location):
    """Save user trip planning data to Excel."""
    try:
        excel_path = Path(settings.USER_DATA_EXCEL)

        if excel_path.exists():
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        else:
            wb = init_excel()
            ws = wb.active

        # Get next row
        next_row = ws.max_row + 1
        s_no = next_row - 1  # Subtract header row

        # Style for data rows
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Alternate row colors
        if s_no % 2 == 0:
            row_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        styles_list = user_data.get('destination_styles', [])
        styles_str = ', '.join(styles_list) if isinstance(styles_list, list) else str(styles_list)

        row_data = [
            s_no,
            user_data.get('name', 'Unknown'),
            ip_address,
            ip_location,
            user_data.get('from_location', 'Not specified'),
            user_data.get('budget', 'Not specified'),
            user_data.get('currency', 'INR'),
            user_data.get('travel_type', 'solo').capitalize(),
            user_data.get('group_size', 1) if user_data.get('travel_type') == 'group' else 'Solo',
            user_data.get('travel_scope', 'within_country').replace('_', ' ').title(),
            user_data.get('num_days', 'Not specified'),
            user_data.get('food_accommodation', 'Not specified').replace('_', ' ').title(),
            user_data.get('travel_medium', 'Not specified').replace('_', ' ').title(),
            styles_str,
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = row_fill

        ws.row_dimensions[next_row].height = 20

        wb.save(excel_path)
        return True
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False
